import pandas as pd
from openpyxl import load_workbook
import os

path_folder = str(os.path.realpath(__file__))[:-7] + "CSV"
file_list = os.listdir(path_folder)


def save_loop(file_name):
    ledger = load_workbook('MYMONEY.xlsx')
    main_df = pd.read_csv(path_folder + "\\" + file_name, names=['date', 'name', 'in', 'out', 'total'])
    main_df[['month', 'day', 'year']] = main_df.date.str.split("/", expand=True,)
    main_df.fillna(0, inplace=True)
    whole_list = []

    for i in range(len(main_df['month'].drop_duplicates())):
        temp_1 = main_df['month'] == list(main_df['month'].drop_duplicates())[i]
        temp_2 = main_df[temp_1]
        temp_3 = temp_2.values.tolist()
        whole_list.append(temp_3)

    for row in range(len(whole_list)):
        for cell in range(len(whole_list[row])):
            if (whole_list[row][cell][7] + '-' + whole_list[row][cell][5]) in ledger.sheetnames:
                origin_sh = ledger[whole_list[row][cell][7] + '-' + whole_list[row][cell][5]]
                origin_sh.append(whole_list[row][cell][0:5])
            else:
                new_sh = ledger.create_sheet(whole_list[row][cell][7] + '-' + whole_list[row][cell][5])
                new_sh.append(['date', 'name', 'in', 'out', '=\'' + whole_list[row][cell][7] + '-'
                               + format(int(whole_list[row][cell][5]) - 1, '02') + '\'!H1', '#####',
                               'money from month', '=LOOKUP(2,1/(E1:E200<>""),E1:E200)'])
                new_sh.append(whole_list[row][cell][0:5])

    for sheet in ledger.sheetnames:
        if sheet == '2021-00':
            continue;
        col_sheet = ledger[sheet]
        for col in range(200):
            if col_sheet['A'+str(col+2)].value == None:
                break;
            cell_name = 'E' + str(col + 2)
            col_sheet[cell_name] = "=E%d-C%d+D%d" % (col + 1, col + 2, col + 2)

    ledger.save('MYMONEY.xlsx')


for file_name in file_list:
    save_loop(file_name)
